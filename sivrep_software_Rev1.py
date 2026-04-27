#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 SIVREP — SISTEMA INSTITUCIONAL DE VALIDACIÓN DE REPUESTOS
         Y LEVANTAMIENTO DE PLANOS ELÉCTRICOS
================================================================================
Versión    : 3.0.0
Autor      : Departamento de Ingeniería — Sistema Institucional
Licencia   : Uso Interno Institucional
Plataforma : Windows / Linux / macOS
Python     : 3.9+

DESCRIPCIÓN GENERAL
───────────────────
Software profesional de escritorio para ingeniería de mantenimiento que permite:
  1. Validación cruzada de maestros de repuestos (Excel / PDF tabulares).
  2. Comparación semántica difusa (fuzzy matching) con ajuste de palabras clave.
  3. Consolidación unificada de listados maestros sin duplicados.
  4. Análisis estadístico completo con gráficos y reportes institucionales.
  5. Extracción OCR de TAGs y Modelos desde planos eléctricos (PDF / Imagen).
  6. Exportación a Excel con formato institucional profesional.

DEPENDENCIAS
────────────
  pip install pandas openpyxl PyPDF2 pdfplumber pytesseract pillow 
              opencv-python-headless numpy matplotlib seaborn 
              fuzzywuzzy python-Levenshtein xlrd PyMuPDF

NOTA OCR
────────
  Requiere Tesseract-OCR instalado en el sistema operativo.
  • Windows: https://github.com/UB-Mannheim/tesseract/wiki
  • Linux : sudo apt install tesseract-ocr tesseract-ocr-spa
  • macOS : brew install tesseract

EJECUCIÓN
─────────
  python sivrep_software.py
================================================================================
"""

import os
import sys
import re
import json
import logging
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional, Any, Callable
import traceback
import time

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN GLOBAL Y CONSTANTES INSTITUCIONALES
# ═══════════════════════════════════════════════════════════════════════════════
@dataclass(frozen=True)
class AppConfig:
    APP_NAME: str = "SIVREP — Validación de Repuestos y Planos Eléctricos"
    VERSION: str = "3.0.0"
    ORGANIZATION: str = "Departamento de Ingeniería"
    DEFAULT_THRESHOLD: int = 75
    MAX_WORDS_MATCH: int = 4
    OUTPUT_DIR: str = "output"
    TEMP_DIR: str = "temp"
    PDF_DPI: int = 300
    OCR_LANG: str = "eng+spa"
    LOGO_TEXT: str = "SIVREP"

CONFIG = AppConfig()
Path(CONFIG.OUTPUT_DIR).mkdir(exist_ok=True)
Path(CONFIG.TEMP_DIR).mkdir(exist_ok=True)

LOG_DIR = Path("logs")
LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = LOG_DIR / f"sivrep_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(module)-25s | %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("SIVREP")

# ═══════════════════════════════════════════════════════════════════════════════
# PALETA DE COLORES PROFESIONAL — TONOS FRÍOS INSTITUCIONALES
# ═══════════════════════════════════════════════════════════════════════════════
class Palette:
    """Paleta cromática institucional. Tonos fríos para entornos de ingeniería."""
    # Fondos
    BG_PRIMARY   = "#f0f4f8"      # Azul grisáceo muy claro (fondo ventana)
    BG_SECONDARY = "#ffffff"      # Blanco puro (tarjetas, paneles)
    BG_TERTIARY  = "#e2e8f0"      # Gris azulado claro (separadores)
    BG_HEADER    = "#1e3a5f"      # Azul marino institucional (encabezados)
    BG_SIDEBAR   = "#2d3748"      # Gris pizarra oscuro (barra lateral)

    # Primarios
    PRIMARY      = "#2b6cb0"      # Azul acero
    PRIMARY_HOVER= "#2c5282"      # Azul acero oscuro
    PRIMARY_LIGHT= "#bee3f8"      # Azul hielo

    # Acentos
    ACCENT       = "#3182ce"      # Azul brillante
    ACCENT_SOFT  = "#63b3ed"      # Azul cielo

    # Texto
    TEXT_PRIMARY   = "#1a202c"    # Casi negro
    TEXT_SECONDARY = "#4a5568"    # Gris oscuro
    TEXT_MUTED     = "#718096"    # Gris medio
    TEXT_ON_DARK   = "#f7fafc"    # Blanco hueso

    # Estados
    SUCCESS      = "#276749"      # Verde bosque frío
    SUCCESS_BG   = "#c6f6d5"      # Verde menta muy claro
    WARNING      = "#c05621"      # Naranja quemado
    WARNING_BG   = "#feebc8"      # Crema cálido
    DANGER       = "#c53030"      # Rojo frío
    DANGER_BG    = "#fed7d7"      # Rosa pálido
    INFO         = "#2b6cb0"      # Azul informativo
    INFO_BG      = "#bee3f8"      # Azul hielo

    # Bordes
    BORDER       = "#cbd5e0"      # Gris borde
    BORDER_FOCUS = "#3182ce"      # Azul foco


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTACIÓN DE LIBRERÍAS DE TERCEROS CON MANEJO GRACEFUL
# ═══════════════════════════════════════════════════════════════════════════════
PANDAS_OK = False
PDF_OK = False
PIL_OK = False
TESSERACT_OK = False
CV2_OK = False
FUZZY_OK = False
MATPLOTLIB_OK = False
SEABORN_OK = False
PYMUPDF_OK = False

try:
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    PANDAS_OK = True
except ImportError as e:
    logger.error(f"[DEP] pandas/openpyxl no disponibles: {e}")

try:
    import PyPDF2
    import pdfplumber
    PDF_OK = True
except ImportError as e:
    logger.error(f"[DEP] PyPDF2/pdfplumber no disponibles: {e}")

try:
    import fitz  # PyMuPDF
    PYMUPDF_OK = True
except ImportError:
    logger.warning("[DEP] PyMuPDF (fitz) no instalado. Conversión PDF→Imagen requerirá pdf2image+poppler.")

try:
    from PIL import Image, ImageTk, ImageEnhance, ImageFilter
    PIL_OK = True
except ImportError as e:
    logger.error(f"[DEP] Pillow no disponible: {e}")

try:
    import pytesseract
    TESSERACT_OK = True
except ImportError as e:
    logger.error(f"[DEP] pytesseract no disponible: {e}")

try:
    import cv2
    CV2_OK = True
except ImportError as e:
    logger.error(f"[DEP] OpenCV no disponible: {e}")

try:
    from fuzzywuzzy import fuzz
    FUZZY_OK = True
except ImportError as e:
    logger.error(f"[DEP] fuzzywuzzy no disponible: {e}")

try:
    import matplotlib
    matplotlib.use("TkAgg")
    import matplotlib.pyplot as plt
    MATPLOTLIB_OK = True
except ImportError as e:
    logger.error(f"[DEP] matplotlib no disponibles: {e}")

try:
    import seaborn as sns
    SEABORN_OK = True
except ImportError:
    pass


# ═══════════════════════════════════════════════════════════════════════════════
# UTILIDADES GENERALES
# ═══════════════════════════════════════════════════════════════════════════════
class Utils:
    """Herramientas utilitarias estáticas para procesamiento de texto técnico."""

    @staticmethod
    def limpiar_texto(texto) -> str:
        if texto is None or (isinstance(texto, float) and pd.isna(texto)):
            return ""
        t = str(texto).upper().strip()
        t = re.sub(r"\s+", " ", t)
        t = re.sub(r"[^A-Z0-9\-_/\.\s]", "", t)
        return t

    @staticmethod
    def extraer_palabras_clave(texto: str, max_palabras: int = 4) -> List[str]:
        texto = Utils.limpiar_texto(texto)
        tokens = [t for t in texto.split() if len(t) > 2 or any(c.isdigit() for c in t)]
        def peso(t):
            return (any(c.isdigit() for c in t), "-" in t, len(t))
        return sorted(tokens, key=peso, reverse=True)[:max_palabras]

    @staticmethod
    def normalizar_codigo_parte(texto: str) -> str:
        texto = Utils.limpiar_texto(texto)
        patrones = [
            r"[A-Z]{2,6}[-\d]{3,20}[A-Z0-9\-]*",
            r"\d{3,6}[A-Z]{1,4}\d{0,4}",
            r"[A-Z]\d{2,4}[A-Z]\d{2,4}",
        ]
        for p in patrones:
            m = re.search(p, texto)
            if m:
                return m.group(0)
        return texto

    @staticmethod
    def calcular_similitud(s1: str, s2: str, metodo: str = "token_sort") -> float:
        if not FUZZY_OK:
            set1, set2 = set(s1.split()), set(s2.split())
            if not set1 or not set2:
                return 0.0
            return (len(set1 & set2) / len(set1 | set2)) * 100
        s1 = Utils.limpiar_texto(s1)
        s2 = Utils.limpiar_texto(s2)
        if metodo == "ratio":
            return float(fuzz.ratio(s1, s2))
        elif metodo == "partial":
            return float(fuzz.partial_ratio(s1, s2))
        elif metodo == "token_sort":
            return float(fuzz.token_sort_ratio(s1, s2))
        elif metodo == "token_set":
            return float(fuzz.token_set_ratio(s1, s2))
        elif metodo == "weighted":
            r1 = fuzz.token_set_ratio(s1, s2)
            r2 = fuzz.partial_ratio(s1, s2)
            r3 = fuzz.ratio(s1, s2)
            return r1 * 0.5 + r2 * 0.3 + r3 * 0.2
        return float(fuzz.token_sort_ratio(s1, s2))

    @staticmethod
    def es_codigo_parte(texto: str) -> bool:
        t = Utils.limpiar_texto(texto)
        if len(t) < 5:
            return False
        tiene_num = any(c.isdigit() for c in t)
        ratio = sum(1 for c in t if c.isdigit()) / len(t)
        return tiene_num and ("-" in t or ratio > 0.3)

    @staticmethod
    def ahora() -> str:
        return datetime.now().strftime("%d/%m/%Y %H:%M:%S")


# ═══════════════════════════════════════════════════════════════════════════════
# MOTOR DE LECTURA DE ARCHIVOS (EXCEL / PDF TABULAR)
# ═══════════════════════════════════════════════════════════════════════════════
class FileReaderEngine:
    """Motor unificado de lectura de archivos maestros."""

    def __init__(self):
        self.historial: List[Dict] = []
        logger.info("FileReaderEngine inicializado.")

    def read_excel(self, ruta: str, hoja: Optional[str] = None) -> "pd.DataFrame":
        if not PANDAS_OK:
            raise RuntimeError("Pandas no está instalado.")
        if not Path(ruta).exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")
        logger.info(f"Leyendo Excel: {ruta}")
        if hoja:
            df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
        else:
            xl = pd.ExcelFile(ruta, engine="openpyxl")
            hoja = xl.sheet_names[0]
            df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
        df = self._limpiar_df(df)
        self.historial.append({"archivo": ruta, "tipo": "excel", "filas": len(df), "columnas": len(df.columns)})
        logger.info(f"Excel cargado: {len(df)} filas × {len(df.columns)} columnas")
        return df

    def read_pdf_tablas(self, ruta: str) -> Optional["pd.DataFrame"]:
        if not PDF_OK:
            raise RuntimeError("Librerías PDF no disponibles.")
        logger.info(f"Extrayendo tablas de PDF: {ruta}")
        tablas = []
        with pdfplumber.open(ruta) as pdf:
            for i, pagina in enumerate(pdf.pages):
                for tabla in pagina.extract_tables() or []:
                    if tabla and len(tabla) > 1:
                        tablas.append(pd.DataFrame(tabla[1:], columns=tabla[0]))
        if tablas:
            df = pd.concat(tablas, ignore_index=True)
            df = self._limpiar_df(df)
            self.historial.append({"archivo": ruta, "tipo": "pdf_tabla", "filas": len(df)})
            return df
        logger.warning("No se detectaron tablas en el PDF.")
        return None

    def read_pdf_texto(self, ruta: str) -> str:
        if not PDF_OK:
            raise RuntimeError("Librerías PDF no disponibles.")
        partes = []
        try:
            with pdfplumber.open(ruta) as pdf:
                for p in pdf.pages:
                    txt = p.extract_text()
                    if txt:
                        partes.append(txt)
        except Exception:
            with open(ruta, "rb") as f:
                lector = PyPDF2.PdfReader(f)
                for p in lector.pages:
                    partes.append(p.extract_text() or "")
        return "\n".join(partes)

    def pdf_a_imagenes(self, ruta: str, dpi: int = 300) -> List[str]:
        """Convierte PDF a imágenes usando PyMuPDF (sin Poppler)."""
        if not PIL_OK:
            raise RuntimeError("Pillow no disponible.")
        rutas_salida = []
        if PYMUPDF_OK:
            logger.info(f"Convirtiendo PDF con PyMuPDF: {ruta}")
            doc = fitz.open(ruta)
            zoom = dpi / 72.0
            mat = fitz.Matrix(zoom, zoom)
            for i in range(len(doc)):
                pagina = doc.load_page(i)
                pix = pagina.get_pixmap(matrix=mat)
                out = Path(CONFIG.TEMP_DIR) / f"pdf_page_{i:03d}.png"
                pix.save(str(out))
                rutas_salida.append(str(out))
            doc.close()
            return rutas_salida
        else:
            # Fallback con advertencia explícita
            raise RuntimeError(
                "Para procesar PDFs como imágenes instale PyMuPDF:\n"
                "  pip install PyMuPDF\n"
                "Alternativa: convierta el PDF a imágenes manualmente y use la opción de imagen."
            )

    def _limpiar_df(self, df: "pd.DataFrame") -> "pd.DataFrame":
        df = df.dropna(how="all")
        df = df.dropna(axis=1, how="all")
        df.columns = [str(c).strip().upper() for c in df.columns]
        return df.reset_index(drop=True)


# ═══════════════════════════════════════════════════════════════════════════════
# MOTOR OCR — OPTIMIZADO PARA PLANOS ELÉCTRICOS INDUSTRIALES
# ═══════════════════════════════════════════════════════════════════════════════
class OCREngine:
    """Extracción inteligente de TAGs y Modelos desde planos eléctricos."""

    def __init__(self):
        self.config_tesseract = f"--psm 6 -l {CONFIG.OCR_LANG}"
        self.patrones = [
            ("TAG_EXPLICITO", r"TAG\s*[:\-]?\s*([A-Z0-9\-]{3,20})"),
            ("TAG_GENERICO", r"([A-Z]{1,4}[-\.]?\d{2,5}[A-Z0-9\-]*)"),
            ("NPARTE_EXPLICITO", r"N[°\s]*PARTE\s*[:\-]?\s*([A-Z0-9\-]{3,25})"),
            ("PARTNO", r"PART\s*NO\s*[:\-]?\s*([A-Z0-9\-]{3,25})"),
            ("MODELO_EXPLICITO", r"MODELO\s*[:\-]?\s*([A-Z0-9\-\s]{3,30})"),
            ("MODEL_EXPLICITO", r"MODEL\s*[:\-]?\s*([A-Z0-9\-\s]{3,30})"),
            ("SERIE", r"SERIE\s*[:\-]?\s*([A-Z0-9\-]{3,20})"),
            ("BANNER_SENSOR", r"S18[A-Z0-9]{5,15}"),
            ("CODIGO_3LETRAS", r"\d{3}[A-Z]{1,4}\d{0,4}"),
        ]
        logger.info("OCREngine inicializado.")

    def preprocesar(self, ruta_imagen: str) -> "np.ndarray":
        if not CV2_OK:
            raise RuntimeError("OpenCV no disponible.")
        img = cv2.imread(ruta_imagen, cv2.IMREAD_GRAYSCALE)
        if img is None:
            raise ValueError(f"No se pudo cargar: {ruta_imagen}")
        img = cv2.resize(img, None, fx=1.5, fy=1.5, interpolation=cv2.INTER_CUBIC)
        img = cv2.fastNlMeansDenoising(img, None, 10, 7, 21)
        img = cv2.adaptiveThreshold(img, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                     cv2.THRESH_BINARY, 11, 2)
        kernel = np.ones((1, 1), np.uint8)
        img = cv2.dilate(img, kernel, iterations=1)
        return img

    def extraer_texto(self, ruta_imagen: str, preprocesar: bool = True) -> str:
        if not TESSERACT_OK:
            raise RuntimeError("Tesseract OCR no disponible.")
        if preprocesar and CV2_OK:
            img = self.preprocesar(ruta_imagen)
            temp = Path(CONFIG.TEMP_DIR) / f"ocr_preproc_{Path(ruta_imagen).stem}.png"
            cv2.imwrite(str(temp), img)
            return pytesseract.image_to_string(img, config=self.config_tesseract).upper()
        return pytesseract.image_to_string(Image.open(ruta_imagen), config=self.config_tesseract).upper()

    def extraer_datos(self, texto: str) -> List[Dict[str, str]]:
        resultados = []
        lineas = [l.strip() for l in texto.split("\n") if len(l.strip()) > 2]
        for i, linea in enumerate(lineas):
            tag = None
            modelo = None
            # Buscar TAG
            for nombre, patron in self.patrones[:2]:
                m = re.search(patron, linea)
                if m:
                    candidato = m.group(1) if m.groups() else m.group(0)
                    if len(candidato) > 2:
                        tag = candidato
                        break
            # Buscar modelo en línea actual
            for nombre, patron in self.patrones[2:]:
                m = re.search(patron, linea)
                if m:
                    candidato = m.group(1) if m.groups() else m.group(0)
                    if Utils.es_codigo_parte(candidato):
                        modelo = candidato
                        break
            # Buscar modelo en línea siguiente
            if tag and not modelo and i + 1 < len(lineas):
                for nombre, patron in self.patrones[2:]:
                    m = re.search(patron, lineas[i + 1])
                    if m:
                        candidato = m.group(1) if m.groups() else m.group(0)
                        if Utils.es_codigo_parte(candidato):
                            modelo = candidato
                            break
            if tag or modelo:
                resultados.append({
                    "TAG": tag or "NO_IDENTIFICADO",
                    "MODELO": modelo or "NO_IDENTIFICADO",
                    "TEXTO_ORIGINAL": linea,
                    "METODO_EXTRACCION": "OCR_ESTRUCTURADO"
                })
        if not resultados:
            resultados = self._extraer_por_bloques(texto)
        return resultados

    def _extraer_por_bloques(self, texto: str) -> List[Dict[str, str]]:
        resultados = []
        palabras = texto.split()
        i = 0
        while i < len(palabras) - 1:
            w1, w2 = palabras[i], palabras[i + 1] if i + 1 < len(palabras) else ""
            if re.match(r"^[A-Z]{1,4}[-\.]?\d{2,5}", w1) and Utils.es_codigo_parte(w2):
                resultados.append({"TAG": w1, "MODELO": w2, "TEXTO_ORIGINAL": f"{w1} {w2}", "METODO_EXTRACCION": "OCR_BLOQUE"})
                i += 2
                continue
            if Utils.es_codigo_parte(w1) and len(w1) > 5:
                resultados.append({"TAG": w1, "MODELO": w2 if Utils.es_codigo_parte(w2) else "NO_IDENTIFICADO",
                                   "TEXTO_ORIGINAL": w1, "METODO_EXTRACCION": "OCR_AISLADO"})
            i += 1
        return resultados

    def procesar_imagen(self, ruta: str) -> List[Dict[str, str]]:
        logger.info(f"OCR sobre imagen: {ruta}")
        texto = self.extraer_texto(ruta)
        return self.extraer_datos(texto)

    def procesar_pdf(self, ruta: str) -> List[Dict[str, str]]:
        logger.info(f"OCR sobre PDF: {ruta}")
        motor = FileReaderEngine()
        imagenes = motor.pdf_a_imagenes(ruta, dpi=CONFIG.PDF_DPI)
        todos = []
        for idx, img in enumerate(imagenes):
            try:
                res = self.procesar_imagen(img)
                for r in res:
                    r["PAGINA_PDF"] = idx + 1
                    r["ARCHIVO_ORIGEN"] = Path(ruta).name
                todos.extend(res)
            except Exception as e:
                logger.error(f"Error página {idx + 1}: {e}")
        logger.info(f"OCR PDF finalizado: {len(todos)} componentes.")
        return todos


# ═══════════════════════════════════════════════════════════════════════════════
# MOTOR DE COMPARACIÓN FUZZY DE REPUESTOS
# ═══════════════════════════════════════════════════════════════════════════════
class SparePartMatcher:
    """Comparación inteligente con múltiples estrategias de matching."""

    def __init__(self, umbral: int = 75, max_palabras: int = 4, estrategia: str = "weighted"):
        self.umbral = umbral
        self.max_palabras = max_palabras
        self.estrategia = estrategia
        self.estadisticas = {
            "total_comparaciones": 0,
            "exactas": 0,
            "fuzzy": 0,
            "sin_coincidencia": 0,
            "tiempo_seg": 0.0
        }
        logger.info(f"Matcher listo (umbral={umbral}, estrategia={estrategia}).")

    def comparar_uno(self, origen: str, destino: str) -> Tuple[bool, float, str]:
        self.estadisticas["total_comparaciones"] += 1
        s = Utils.limpiar_texto(origen)
        t = Utils.limpiar_texto(destino)
        if not s or not t:
            return False, 0.0, "vacio"
        if s == t:
            self.estadisticas["exactas"] += 1
            return True, 100.0, "exacto"
        cs = Utils.normalizar_codigo_parte(s)
        ct = Utils.normalizar_codigo_parte(t)
        if cs and ct and cs == ct:
            self.estadisticas["exactas"] += 1
            return True, 100.0, "codigo_parte_exacto"
        ws = set(Utils.extraer_palabras_clave(s, self.max_palabras))
        wt = set(Utils.extraer_palabras_clave(t, self.max_palabras))
        if ws and wt:
            inter = ws & wt
            if len(inter) >= min(2, len(ws), len(wt)):
                score = (len(inter) / max(len(ws), len(wt))) * 100
                if score >= self.umbral:
                    self.estadisticas["fuzzy"] += 1
                    return True, round(score, 2), "palabras_clave"
        score = Utils.calcular_similitud(s, t, self.estrategia)
        if score >= self.umbral:
            self.estadisticas["fuzzy"] += 1
            return True, round(score, 2), f"fuzzy_{self.estrategia}"
        self.estadisticas["sin_coincidencia"] += 1
        return False, round(score, 2), "sin_coincidencia"

    def comparar_dataframes(self, df_origen: "pd.DataFrame", col_origen: str,
                            df_destino: "pd.DataFrame", col_destino: str,
                            callback_progreso: Optional[Callable] = None) -> "pd.DataFrame":
        t0 = time.time()
        resultados = []
        total = len(df_origen)
        valores_destino = [(i, Utils.limpiar_texto(str(v))) for i, v in enumerate(df_destino[col_destino].astype(str))]
        for idx, fila in df_origen.iterrows():
            val_origen = str(fila[col_origen]) if pd.notna(fila[col_origen]) else ""
            if not val_origen.strip():
                continue
            best_score = 0.0
            best_match = ""
            best_metodo = ""
            best_idx = -1
            for t_idx, t_val in valores_destino:
                match, score, metodo = self.comparar_uno(val_origen, t_val)
                if score > best_score:
                    best_score = score
                    best_match = t_val
                    best_metodo = metodo
                    best_idx = t_idx
                    if score == 100.0:
                        break
            coincide = best_score >= self.umbral
            fila_res = {
                "IDX_ORIGEN": idx,
                "VALOR_ORIGEN": val_origen,
                "VALOR_DESTINO": best_match,
                "SCORE": best_score,
                "METODO": best_metodo,
                "COINCIDE": "SÍ" if coincide else "NO",
                "IDX_DESTINO": best_idx if coincide else -1
            }
            for c in df_origen.columns:
                fila_res[f"SRC_{c}"] = fila[c]
            resultados.append(fila_res)
            if callback_progreso and idx % 10 == 0:
                callback_progreso(int(idx) + 1, total)
        self.estadisticas["tiempo_seg"] = time.time() - t0
        logger.info(f"Comparación lista en {self.estadisticas['tiempo_seg']:.2f}s")
        return pd.DataFrame(resultados)


# ═══════════════════════════════════════════════════════════════════════════════
# MOTOR DE ANÁLISIS ESTADÍSTICO Y REPORTERÍA
# ═══════════════════════════════════════════════════════════════════════════════
class StatisticsEngine:
    """Genera métricas, gráficos y reportes institucionales."""

    def __init__(self):
        self.reporte: Dict[str, Any] = {}
        logger.info("StatisticsEngine inicializado.")

    def analizar(self, df_resultado: "pd.DataFrame", df_origen: "pd.DataFrame",
                 df_destino: "pd.DataFrame") -> Dict[str, Any]:
        total_o, total_d = len(df_origen), len(df_destino)
        if "COINCIDE" not in df_resultado.columns:
            raise ValueError("DataFrame sin columna COINCIDE")
        coin = df_resultado[df_resultado["COINCIDE"] == "SÍ"]
        no_coin = df_resultado[df_resultado["COINCIDE"] == "NO"]
        metricas = {
            "total_origen": total_o,
            "total_destino": total_d,
            "comparados": len(df_resultado),
            "coincidencias": len(coin),
            "no_coincidencias": len(no_coin),
            "tasa_exito": round(len(coin) / len(df_resultado) * 100, 2) if len(df_resultado) else 0,
            "cobertura": round(len(coin) / total_o * 100, 2) if total_o else 0,
            "score_promedio": round(coin["SCORE"].mean(), 2) if len(coin) else 0,
            "score_min": round(coin["SCORE"].min(), 2) if len(coin) else 0,
            "score_max": round(coin["SCORE"].max(), 2) if len(coin) else 0,
        }
        if "METODO" in df_resultado.columns:
            metricas["metodos"] = df_resultado["METODO"].value_counts().to_dict()
        metricas["vacios_origen"] = {c: int(df_origen[c].isna().sum()) for c in df_origen.columns}
        metricas["vacios_destino"] = {c: int(df_destino[c].isna().sum()) for c in df_destino.columns}
        self.reporte = metricas
        return metricas

    def generar_graficos(self) -> Optional[str]:
        if not MATPLOTLIB_OK or not self.reporte:
            return None
        fig, axes = plt.subplots(2, 2, figsize=(13, 10), facecolor=Palette.BG_PRIMARY)
        fig.suptitle("SIVREP — Análisis Estadístico de Validación", fontsize=16,
                     fontweight="bold", color=Palette.BG_HEADER)
        # 1. Pie chart
        ax1 = axes[0, 0]
        sizes = [self.reporte["coincidencias"], self.reporte["no_coincidencias"]]
        colors = ["#38a169", "#e53e3e"]
        ax1.pie(sizes, labels=["Coincidencias", "No Coincidencias"],
                autopct="%1.1f%%", colors=colors, startangle=90,
                textprops={"fontsize": 11, "color": Palette.TEXT_PRIMARY})
        ax1.set_title("Distribución de Resultados", fontsize=12, color=Palette.TEXT_PRIMARY)
        # 2. Barras horizontales
        ax2 = axes[0, 1]
        metricas = ["tasa_exito", "cobertura", "score_promedio"]
        valores = [self.reporte.get(m, 0) for m in metricas]
        nombres = ["Tasa Éxito (%)", "Cobertura (%)", "Score Promedio"]
        bars = ax2.barh(nombres, valores, color=["#3182ce", "#805ad5", "#dd6b20"])
        ax2.set_xlim(0, 100)
        ax2.set_title("Métricas de Desempeño", fontsize=12, color=Palette.TEXT_PRIMARY)
        ax2.tick_params(colors=Palette.TEXT_PRIMARY)
        for bar, val in zip(bars, valores):
            ax2.text(val + 1, bar.get_y() + bar.get_height() / 2, f"{val:.1f}",
                     va="center", fontsize=10, color=Palette.TEXT_PRIMARY)
        # 3. Métodos
        ax3 = axes[1, 0]
        if "metodos" in self.reporte:
            met = self.reporte["metodos"]
            ax3.bar(met.keys(), met.values(), color="#2b6cb0")
            ax3.set_title("Métodos de Coincidencia", fontsize=12, color=Palette.TEXT_PRIMARY)
            ax3.tick_params(axis="x", rotation=45, colors=Palette.TEXT_PRIMARY)
            ax3.tick_params(axis="y", colors=Palette.TEXT_PRIMARY)
        # 4. Resumen
        ax4 = axes[1, 1]
        ax4.axis("off")
        txt = (f"RESUMEN EJECUTIVO\n{'─'*40}\n"
               f"Total Origen:      {self.reporte['total_origen']:,}\n"
               f"Total Destino:     {self.reporte['total_destino']:,}\n"
               f"Comparados:        {self.reporte['comparados']:,}\n"
               f"Coincidencias:     {self.reporte['coincidencias']:,}\n"
               f"No Coincidencias:  {self.reporte['no_coincidencias']:,}\n"
               f"Score Mínimo:      {self.reporte['score_min']}\n"
               f"Score Máximo:      {self.reporte['score_max']}")
        ax4.text(0.1, 0.5, txt, fontsize=11, family="monospace",
                 verticalalignment="center", color=Palette.TEXT_PRIMARY)
        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        ruta = Path(CONFIG.OUTPUT_DIR) / f"estadisticas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        plt.savefig(ruta, dpi=150, bbox_inches="tight", facecolor=Palette.BG_PRIMARY)
        plt.close()
        logger.info(f"Gráficos guardados: {ruta}")
        return str(ruta)

    def generar_excel(self, df_resultado: "pd.DataFrame", df_unificado: Optional["pd.DataFrame"] = None) -> str:
        if not PANDAS_OK:
            raise RuntimeError("Pandas no disponible.")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta = Path(CONFIG.OUTPUT_DIR) / f"REPORTE_SIVREP_{ts}.xlsx"
        with pd.ExcelWriter(ruta, engine="openpyxl") as w:
            df_resultado.to_excel(w, sheet_name="RESULTADOS_DETALLADOS", index=False)
            if "COINCIDE" in df_resultado.columns:
                df_resultado[df_resultado["COINCIDE"] == "SÍ"].to_excel(w, sheet_name="COINCIDENCIAS", index=False)
                df_resultado[df_resultado["COINCIDE"] == "NO"].to_excel(w, sheet_name="NO_COINCIDENCIAS", index=False)
            if df_unificado is not None:
                df_unificado.to_excel(w, sheet_name="LISTADO_UNIFICADO", index=False)
            pd.DataFrame([self.reporte]).to_excel(w, sheet_name="ESTADISTICAS", index=False)
        self._formatear_excel(str(ruta))
        logger.info(f"Excel generado: {ruta}")
        return str(ruta)

    def _formatear_excel(self, ruta: str):
        try:
            wb = load_workbook(ruta)
            header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
            header_font = Font(color="f7fafc", bold=True, size=11, name="Segoe UI")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Border(left=Side(style="thin", color="cbd5e0"),
                          right=Side(style="thin", color="cbd5e0"),
                          top=Side(style="thin", color="cbd5e0"),
                          bottom=Side(style="thin", color="cbd5e0"))
            for nombre in wb.sheetnames:
                ws = wb[nombre]
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.border = thin
                for col in ws.columns:
                    max_len = 0
                    letra = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[letra].width = min(max_len + 3, 55)
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
            wb.save(ruta)
        except Exception as e:
            logger.error(f"Error formateando Excel: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# CONSOLIDADOR UNIFICADO DE MAESTROS
# ═══════════════════════════════════════════════════════════════════════════════
class UnifiedConsolidator:
    """Genera un listado maestro único a partir de dos fuentes."""

    COLUMNAS_PRIORITARIAS = [
        "SISTEMA", "EQUIPO Y SUBEQUIPO", "ESPECIALIDAD", "EQUIPO O REPUESTO",
        "ARTICULO", "FAMILIA", "TAG", "FABRICANTE", "MODELO", "SERIE",
        "N° PARTE", "NOMBRE", "DESCRIPCION"
    ]

    def __init__(self):
        logger.info("UnifiedConsolidator inicializado.")

    def consolidar(self, df_o: "pd.DataFrame", df_d: "pd.DataFrame",
                   df_res: "pd.DataFrame", col_o: str, col_d: str) -> "pd.DataFrame":
        idx_coin = df_res[df_res["COINCIDE"] == "SÍ"]["IDX_ORIGEN"].tolist()
        df_o_m = df_o.copy()
        df_o_m["_ORIGEN"] = "ARCHIVO_1"
        df_o_m["_ESTADO"] = df_o_m.index.map(lambda x: "COINCIDE" if x in idx_coin else "UNICO_A1")
        idx_t_usados = df_res[df_res["COINCIDE"] == "SÍ"]["IDX_DESTINO"].dropna().astype(int).unique()
        df_d_u = df_d.copy()
        df_d_u = df_d_u[~df_d_u.index.isin(idx_t_usados)]
        df_d_u["_ORIGEN"] = "ARCHIVO_2"
        df_d_u["_ESTADO"] = "UNICO_A2"
        no1 = self._normalizar_cols(df_o_m)
        no2 = self._normalizar_cols(df_d_u)
        todas = list(set(no1.columns) | set(no2.columns))
        for c in todas:
            if c not in no1.columns:
                no1[c] = ""
            if c not in no2.columns:
                no2[c] = ""
        orden = [c for c in self.COLUMNAS_PRIORITARIAS if c in todas]
        orden += sorted([c for c in todas if c not in self.COLUMNAS_PRIORITARIAS])
        no1, no2 = no1[orden], no2[orden]
        uni = pd.concat([no1, no2], ignore_index=True)
        uni["_FECHA_CONSOLIDACION"] = Utils.ahora()
        uni["_ID_UNICO"] = [f"REP-{i+1:06d}" for i in range(len(uni))]
        logger.info(f"Consolidado: {len(uni)} repuestos únicos.")
        return uni

    def _normalizar_cols(self, df: "pd.DataFrame") -> "pd.DataFrame":
        df = df.copy()
        mapping = {
            r"SISTEMA|SIST": "SISTEMA",
            r"EQUIPO\s*Y\s*SUBEQUIPO|EQUIPO\s*SUB|EQUIPO(?!\s*O)": "EQUIPO Y SUBEQUIPO",
            r"ESPECIALIDAD|ESP": "ESPECIALIDAD",
            r"EQUIPO\s*O\s*REPUESTO|REPUESTO|EQUIPO\s*REP": "EQUIPO O REPUESTO",
            r"ARTICULO|ARTÍCULO|ART(?!I)": "ARTICULO",
            r"FAMILIA|FAM": "FAMILIA",
            r"^TAG$|TAG\s*ID": "TAG",
            r"FABRICANTE|FABR|MARCA|MANUFACTURER|MAKER": "FABRICANTE",
            r"MODELO|MODEL(?!O)|MOD(?=ELO)": "MODELO",
            r"SERIE|SER(?!I)|SERIAL": "SERIE",
            r"N[°\s]*PARTE|PART\s*NO|PART\s*NUMBER|NUMERO\s*PARTE|N\s*PARTE|N°\s*PARTE|PARTE": "N° PARTE",
            r"NOMBRE|DESCRIPCION|DESCRIPTION|DESC(?!R)": "NOMBRE",
        }
        nuevas = {}
        for col in df.columns:
            asignado = False
            for patron, estandar in mapping.items():
                if re.search(patron, col, re.IGNORECASE):
                    nuevas[col] = estandar
                    asignado = True
                    break
            if not asignado:
                nuevas[col] = col
        df.rename(columns=nuevas, inplace=True)
        return df


# ═══════════════════════════════════════════════════════════════════════════════
# INTERFAZ GRÁFICA PROFESIONAL — DISEÑO INSTITUCIONAL FRÍO
# ═══════════════════════════════════════════════════════════════════════════════
class SIVREPApp:
    """Aplicación de escritorio con estética profesional de ingeniería."""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(f"{CONFIG.APP_NAME}  v{CONFIG.VERSION}")
        self.root.geometry("1500x950")
        self.root.minsize(1300, 850)
        self.root.configure(bg=Palette.BG_PRIMARY)

        # Variables
        self.df_origen: Optional["pd.DataFrame"] = None
        self.df_destino: Optional["pd.DataFrame"] = None
        self.df_resultado: Optional["pd.DataFrame"] = None
        self.df_unificado: Optional["pd.DataFrame"] = None
        self.df_ocr: Optional["pd.DataFrame"] = None
        self.var_col_origen = tk.StringVar()
        self.var_col_destino = tk.StringVar()
        self.var_umbral = tk.IntVar(value=CONFIG.DEFAULT_THRESHOLD)
        self.var_max_pal = tk.IntVar(value=CONFIG.MAX_WORDS_MATCH)
        self.var_estrategia = tk.StringVar(value="weighted")

        # Motores
        self.motor_archivos = FileReaderEngine()
        self.motor_ocr = OCREngine()
        self.matcher = SparePartMatcher()
        self.stats = StatisticsEngine()
        self.consolidador = UnifiedConsolidator()

        self._configurar_estilos()
        self._construir_ui()
        self._verificar_dependencias()
        logger.info("SIVREPApp iniciada.")

    # ──────────────────────────────────────────────────────────────────────────
    # CONFIGURACIÓN DE ESTILOS TTK PROFESIONALES
    # ──────────────────────────────────────────────────────────────────────────
    def _configurar_estilos(self):
        style = ttk.Style(self.root)
        style.theme_use("clam")

        # Fuente base profesional
        fuente_base = ("Segoe UI", 10)
        fuente_header = ("Segoe UI", 11, "bold")
        fuente_titulo = ("Segoe UI", 13, "bold")
        fuente_peq = ("Segoe UI", 9)

        # Frame general
        style.configure("TFrame", background=Palette.BG_PRIMARY)
        style.configure("TLabel", background=Palette.BG_PRIMARY, foreground=Palette.TEXT_PRIMARY,
                        font=fuente_base)
        style.configure("TButton", font=fuente_base, padding=6)
        style.configure("TNotebook", background=Palette.BG_PRIMARY, tabmargins=[2, 5, 2, 0])
        style.configure("TNotebook.Tab", font=fuente_header, padding=[12, 4],
                        background=Palette.BG_TERTIARY, foreground=Palette.TEXT_SECONDARY)
        style.map("TNotebook.Tab",
                  background=[("selected", Palette.PRIMARY), ("active", Palette.PRIMARY_LIGHT)],
                  foreground=[("selected", Palette.TEXT_ON_DARK), ("active", Palette.TEXT_PRIMARY)])

        # LabelFrame
        style.configure("TLabelframe", background=Palette.BG_SECONDARY, borderwidth=2,
                        relief="solid", bordercolor=Palette.BORDER)
        style.configure("TLabelframe.Label", background=Palette.BG_SECONDARY,
                        foreground=Palette.PRIMARY, font=fuente_header)

        # Treeview profesional
        style.configure("Custom.Treeview",
                        background=Palette.BG_SECONDARY,
                        foreground=Palette.TEXT_PRIMARY,
                        fieldbackground=Palette.BG_SECONDARY,
                        rowheight=26,
                        font=fuente_base)
        style.configure("Custom.Treeview.Heading",
                        background=Palette.BG_HEADER,
                        foreground=Palette.TEXT_ON_DARK,
                        font=fuente_header,
                        relief="flat")
        style.map("Custom.Treeview.Heading", background=[("active", Palette.PRIMARY_HOVER)])
        style.map("Custom.Treeview",
                  background=[("selected", Palette.PRIMARY_LIGHT)],
                  foreground=[("selected", Palette.TEXT_PRIMARY)])

        # Progressbar
        style.configure("Horizontal.TProgressbar", thickness=18,
                        background=Palette.ACCENT, troughcolor=Palette.BG_TERTIARY,
                        bordercolor=Palette.BORDER, lightcolor=Palette.ACCENT_SOFT)

        # Combobox
        style.configure("TCombobox", font=fuente_base)
        style.map("TCombobox", fieldbackground=[("readonly", Palette.BG_SECONDARY)])

        # Spinbox
        style.configure("TSpinbox", font=fuente_base, fieldbackground=Palette.BG_SECONDARY)

        # Scrollbar
        style.configure("Vertical.TScrollbar", background=Palette.BG_TERTIARY,
                        troughcolor=Palette.BG_PRIMARY, bordercolor=Palette.BORDER,
                        arrowcolor=Palette.TEXT_SECONDARY)
        style.configure("Horizontal.TScrollbar", background=Palette.BG_TERTIARY,
                        troughcolor=Palette.BG_PRIMARY, bordercolor=Palette.BORDER,
                        arrowcolor=Palette.TEXT_SECONDARY)

        # Botones de acción personalizados
        style.configure("Action.TButton", font=fuente_header, foreground=Palette.TEXT_ON_DARK,
                        background=Palette.PRIMARY)
        style.map("Action.TButton",
                  background=[("active", Palette.PRIMARY_HOVER), ("pressed", Palette.BG_HEADER)],
                  foreground=[("active", Palette.TEXT_ON_DARK)])

        style.configure("Success.TButton", font=fuente_header, foreground=Palette.TEXT_ON_DARK,
                        background=Palette.SUCCESS)
        style.map("Success.TButton",
                  background=[("active", "#2f855a"), ("pressed", Palette.SUCCESS)])

    # ──────────────────────────────────────────────────────────────────────────
    # CONSTRUCCIÓN DE LA INTERFAZ
    # ──────────────────────────────────────────────────────────────────────────
    def _construir_ui(self):
        # Barra superior institucional
        self._crear_barra_superior()

        # Notebook principal
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=14, pady=(0, 10))

        # Pestañas
        self.tab_val = tk.Frame(self.notebook, bg=Palette.BG_PRIMARY)
        self.notebook.add(self.tab_val, text="   Validación de Repuestos   ")
        self._construir_tab_validacion()

        self.tab_plan = tk.Frame(self.notebook, bg=Palette.BG_PRIMARY)
        self.notebook.add(self.tab_plan, text="   Planos Eléctricos (OCR)   ")
        self._construir_tab_planos()

        self.tab_est = tk.Frame(self.notebook, bg=Palette.BG_PRIMARY)
        self.notebook.add(self.tab_est, text="   Estadísticas y Reportes   ")
        self._construir_tab_estadisticas()

        # Barra de estado inferior
        self.barra_estado = tk.Label(self.root, text="  Sistema listo  ",
                                     bg=Palette.BG_HEADER, fg=Palette.TEXT_ON_DARK,
                                     font=("Segoe UI", 9), anchor="w", padx=10, pady=4)
        self.barra_estado.pack(side="bottom", fill="x")

    def _crear_barra_superior(self):
        barra = tk.Frame(self.root, bg=Palette.BG_HEADER, height=56)
        barra.pack(fill="x", side="top")
        barra.pack_propagate(False)

        lbl_logo = tk.Label(barra, text=CONFIG.LOGO_TEXT,
                            bg=Palette.BG_HEADER, fg=Palette.TEXT_ON_DARK,
                            font=("Segoe UI", 18, "bold"))
        lbl_logo.pack(side="left", padx=(20, 10), pady=8)

        lbl_sub = tk.Label(barra, text=f"v{CONFIG.VERSION}  |  {CONFIG.ORGANIZATION}",
                           bg=Palette.BG_HEADER, fg="#a0aec0",
                           font=("Segoe UI", 10))
        lbl_sub.pack(side="left", pady=8)

        lbl_fecha = tk.Label(barra, text=Utils.ahora().split()[0],
                             bg=Palette.BG_HEADER, fg="#a0aec0",
                             font=("Segoe UI", 10))
        lbl_fecha.pack(side="right", padx=20, pady=8)

    def _construir_tab_validacion(self):
        # Panel izquierdo: carga y configuración
        panel_izq = tk.Frame(self.tab_val, bg=Palette.BG_PRIMARY)
        panel_izq.pack(side="left", fill="y", padx=(10, 5), pady=10)

        # ── Sección Archivos ──
        frm_arch = tk.LabelFrame(panel_izq, text="  Carga de Archivos  ",
                                 bg=Palette.BG_SECONDARY, fg=Palette.PRIMARY,
                                 font=("Segoe UI", 11, "bold"),
                                 padx=12, pady=10, bd=2, relief="solid")
        frm_arch.pack(fill="x", pady=(0, 10))
        frm_arch.configure(highlightbackground=Palette.BORDER, highlightthickness=1)

        self._fila_archivo(frm_arch, 0, "Archivo 1 (Base):", "_cargar_origen",
                           "lbl_archivo_origen", Palette.SUCCESS)
        self._fila_archivo(frm_arch, 1, "Archivo 2 (Comparar):", "_cargar_destino",
                           "lbl_archivo_destino", Palette.SUCCESS)

        # ── Sección Configuración ──
        frm_cfg = tk.LabelFrame(panel_izq, text="  Parámetros de Comparación  ",
                                bg=Palette.BG_SECONDARY, fg=Palette.PRIMARY,
                                font=("Segoe UI", 11, "bold"),
                                padx=12, pady=10, bd=2, relief="solid")
        frm_cfg.pack(fill="x", pady=(0, 10))
        frm_cfg.configure(highlightbackground=Palette.BORDER, highlightthickness=1)

        tk.Label(frm_cfg, text="Columna Archivo 1:", bg=Palette.BG_SECONDARY,
                 fg=Palette.TEXT_SECONDARY, font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w", pady=6)
        self.cmb_origen = ttk.Combobox(frm_cfg, textvariable=self.var_col_origen,
                                       state="readonly", width=32, font=("Segoe UI", 10))
        self.cmb_origen.grid(row=0, column=1, padx=8, pady=6)

        tk.Label(frm_cfg, text="Columna Archivo 2:", bg=Palette.BG_SECONDARY,
                 fg=Palette.TEXT_SECONDARY, font=("Segoe UI", 10)).grid(row=1, column=0, sticky="w", pady=6)
        self.cmb_destino = ttk.Combobox(frm_cfg, textvariable=self.var_col_destino,
                                        state="readonly", width=32, font=("Segoe UI", 10))
        self.cmb_destino.grid(row=1, column=1, padx=8, pady=6)

        tk.Label(frm_cfg, text="Umbral mínimo (%):", bg=Palette.BG_SECONDARY,
                 fg=Palette.TEXT_SECONDARY, font=("Segoe UI", 10)).grid(row=2, column=0, sticky="w", pady=6)
        ttk.Spinbox(frm_cfg, from_=0, to=100, textvariable=self.var_umbral,
                    width=12, font=("Segoe UI", 10)).grid(row=2, column=1, sticky="w", padx=8, pady=6)

        tk.Label(frm_cfg, text="Máx. palabras clave:", bg=Palette.BG_SECONDARY,
                 fg=Palette.TEXT_SECONDARY, font=("Segoe UI", 10)).grid(row=3, column=0, sticky="w", pady=6)
        ttk.Spinbox(frm_cfg, from_=1, to=10, textvariable=self.var_max_pal,
                    width=12, font=("Segoe UI", 10)).grid(row=3, column=1, sticky="w", padx=8, pady=6)

        tk.Label(frm_cfg, text="Estrategia fuzzy:", bg=Palette.BG_SECONDARY,
                 fg=Palette.TEXT_SECONDARY, font=("Segoe UI", 10)).grid(row=4, column=0, sticky="w", pady=6)
        ttk.Combobox(frm_cfg, textvariable=self.var_estrategia,
                     values=["ratio", "partial", "token_sort", "token_set", "weighted"],
                     state="readonly", width=20, font=("Segoe UI", 10)).grid(row=4, column=1, sticky="w", padx=8, pady=6)

        # ── Botones de acción ──
        frm_btn = tk.Frame(panel_izq, bg=Palette.BG_PRIMARY)
        frm_btn.pack(fill="x", pady=(5, 10))

        tk.Button(frm_btn, text="▶  EJECUTAR VALIDACIÓN", command=self._ejecutar_validacion,
                  bg=Palette.PRIMARY, fg=Palette.TEXT_ON_DARK, font=("Segoe UI", 11, "bold"),
                  activebackground=Palette.PRIMARY_HOVER, activeforeground=Palette.TEXT_ON_DARK,
                  cursor="hand2", relief="flat", padx=14, pady=8).pack(fill="x", pady=3)

        tk.Button(frm_btn, text="💾  Exportar Resultados", command=self._exportar_resultados,
                  bg=Palette.BG_TERTIARY, fg=Palette.TEXT_PRIMARY, font=("Segoe UI", 10, "bold"),
                  activebackground=Palette.BORDER, cursor="hand2", relief="flat", padx=10, pady=6).pack(fill="x", pady=3)

        tk.Button(frm_btn, text="📊  Excel Unificado", command=self._exportar_unificado,
                  bg=Palette.SUCCESS_BG, fg=Palette.SUCCESS, font=("Segoe UI", 10, "bold"),
                  activebackground="#9ae6b4", cursor="hand2", relief="flat", padx=10, pady=6).pack(fill="x", pady=3)

        # Barra de progreso
        self.progreso = ttk.Progressbar(panel_izq, mode="determinate", length=280)
        self.progreso.pack(fill="x", pady=(8, 0))

        # ── Panel derecho: resultados en Notebook ──
        panel_der = tk.Frame(self.tab_val, bg=Palette.BG_PRIMARY)
        panel_der.pack(side="left", fill="both", expand=True, padx=(5, 10), pady=10)

        self.nb_resultados = ttk.Notebook(panel_der)
        self.nb_resultados.pack(fill="both", expand=True)

        self.tree_vista_origen = self._crear_tree_en_notebook(self.nb_resultados, "Vista Archivo 1")
        self.tree_vista_destino = self._crear_tree_en_notebook(self.nb_resultados, "Vista Archivo 2")
        self.tree_resultados = self._crear_tree_en_notebook(self.nb_resultados, "Resultados Comparación")
        self.tree_no_match = self._crear_tree_en_notebook(self.nb_resultados, "No Coincidencias")

    def _fila_archivo(self, parent, row, etiqueta, comando, attr_lbl, color_ok):
        tk.Label(parent, text=etiqueta, bg=Palette.BG_SECONDARY,
                 fg=Palette.TEXT_SECONDARY, font=("Segoe UI", 10), width=20, anchor="w").grid(row=row, column=0, sticky="w", pady=5)
        lbl = tk.Label(parent, text="No seleccionado", bg=Palette.BG_SECONDARY,
                       fg=Palette.TEXT_MUTED, font=("Segoe UI", 10, "italic"), anchor="w")
        lbl.grid(row=row, column=1, sticky="ew", padx=5, pady=5)
        setattr(self, attr_lbl, lbl)
        tk.Button(parent, text="Examinar…", command=getattr(self, comando),
                  bg=Palette.BG_TERTIARY, fg=Palette.TEXT_PRIMARY, font=("Segoe UI", 9),
                  activebackground=Palette.BORDER, relief="flat", cursor="hand2",
                  padx=8, pady=2).grid(row=row, column=2, padx=5, pady=5)

    def _construir_tab_planos(self):
        # Panel superior de carga
        frm_top = tk.LabelFrame(self.tab_plan, text="  Carga de Planos Eléctricos  ",
                                bg=Palette.BG_SECONDARY, fg=Palette.PRIMARY,
                                font=("Segoe UI", 11, "bold"),
                                padx=14, pady=10, bd=2, relief="solid")
        frm_top.pack(fill="x", padx=10, pady=(10, 5))
        frm_top.configure(highlightbackground=Palette.BORDER, highlightthickness=1)

        tk.Label(frm_top, text="Archivo (PDF o Imagen):", bg=Palette.BG_SECONDARY,
                 fg=Palette.TEXT_SECONDARY, font=("Segoe UI", 10), width=22, anchor="w").grid(row=0, column=0, sticky="w")
        self.lbl_plano = tk.Label(frm_top, text="No seleccionado", bg=Palette.BG_SECONDARY,
                                  fg=Palette.TEXT_MUTED, font=("Segoe UI", 10, "italic"), anchor="w")
        self.lbl_plano.grid(row=0, column=1, sticky="ew", padx=6, pady=6)
        tk.Button(frm_top, text="Examinar…", command=self._cargar_plano,
                  bg=Palette.BG_TERTIARY, fg=Palette.TEXT_PRIMARY, font=("Segoe UI", 9),
                  activebackground=Palette.BORDER, relief="flat", cursor="hand2",
                  padx=10, pady=3).grid(row=0, column=2, padx=6, pady=6)
        tk.Button(frm_top, text="▶  PROCESAR OCR", command=self._ejecutar_ocr,
                  bg=Palette.PRIMARY, fg=Palette.TEXT_ON_DARK, font=("Segoe UI", 11, "bold"),
                  activebackground=Palette.PRIMARY_HOVER, relief="flat", cursor="hand2",
                  padx=16, pady=6).grid(row=0, column=3, padx=(20, 6), pady=6)
        tk.Button(frm_top, text="💾  Exportar Excel", command=self._exportar_ocr,
                  bg=Palette.SUCCESS_BG, fg=Palette.SUCCESS, font=("Segoe UI", 10, "bold"),
                  activebackground="#9ae6b4", relief="flat", cursor="hand2",
                  padx=12, pady=6).grid(row=0, column=4, padx=6, pady=6)

        # Panel dividido: Tree arriba, Log abajo
        panel_medio = tk.PanedWindow(self.tab_plan, orient="vertical",
                                     bg=Palette.BORDER, sashwidth=4)
        panel_medio.pack(fill="both", expand=True, padx=10, pady=5)

        # Tree de resultados OCR
        frm_tree = tk.LabelFrame(panel_medio, text="  Componentes Detectados  ",
                                 bg=Palette.BG_SECONDARY, fg=Palette.PRIMARY,
                                 font=("Segoe UI", 11, "bold"),
                                 padx=6, pady=6, bd=2, relief="solid")
        frm_tree.configure(highlightbackground=Palette.BORDER, highlightthickness=1)
        panel_medio.add(frm_tree, minsize=300)

        self.tree_ocr = ttk.Treeview(frm_tree, style="Custom.Treeview")
        vsb = ttk.Scrollbar(frm_tree, orient="vertical", command=self.tree_ocr.yview)
        hsb = ttk.Scrollbar(frm_tree, orient="horizontal", command=self.tree_ocr.xview)
        self.tree_ocr.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree_ocr.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frm_tree.grid_rowconfigure(0, weight=1)
        frm_tree.grid_columnconfigure(0, weight=1)

        # Log de OCR
        frm_log = tk.LabelFrame(panel_medio, text="  Log de Procesamiento OCR  ",
                                bg=Palette.BG_SECONDARY, fg=Palette.PRIMARY,
                                font=("Segoe UI", 11, "bold"),
                                padx=6, pady=6, bd=2, relief="solid")
        frm_log.configure(highlightbackground=Palette.BORDER, highlightthickness=1)
        panel_medio.add(frm_log, minsize=180)

        self.txt_log_ocr = scrolledtext.ScrolledText(frm_log, wrap=tk.WORD,
                                                      font=("Consolas", 9),
                                                      bg="#f7fafc", fg=Palette.TEXT_PRIMARY,
                                                      insertbackground=Palette.TEXT_PRIMARY,
                                                      relief="flat", padx=6, pady=6)
        self.txt_log_ocr.pack(fill="both", expand=True)
        self.txt_log_ocr.insert(tk.END, "[SISTEMA] Listo para procesar planos eléctricos.\n")
        self.txt_log_ocr.config(state="disabled")

    def _construir_tab_estadisticas(self):
        frm_top = tk.Frame(self.tab_est, bg=Palette.BG_PRIMARY)
        frm_top.pack(fill="x", padx=10, pady=(10, 5))

        botones = [
            ("📈  Generar Análisis", self._generar_estadisticas, Palette.PRIMARY),
            ("🖼  Ver Gráficos", self._ver_graficos, Palette.ACCENT),
            ("📋  Reporte JSON", self._ver_json, Palette.BG_SIDEBAR),
        ]
        for texto, cmd, color in botones:
            tk.Button(frm_top, text=texto, command=cmd,
                      bg=color, fg=Palette.TEXT_ON_DARK, font=("Segoe UI", 10, "bold"),
                      activebackground=Palette.PRIMARY_HOVER, relief="flat",
                      cursor="hand2", padx=14, pady=7).pack(side="left", padx=5)

        frm_txt = tk.LabelFrame(self.tab_est, text="  Resumen Estadístico Institucional  ",
                                bg=Palette.BG_SECONDARY, fg=Palette.PRIMARY,
                                font=("Segoe UI", 11, "bold"),
                                padx=8, pady=8, bd=2, relief="solid")
        frm_txt.pack(fill="both", expand=True, padx=10, pady=5)
        frm_txt.configure(highlightbackground=Palette.BORDER, highlightthickness=1)

        self.txt_estadisticas = scrolledtext.ScrolledText(frm_txt, wrap=tk.WORD,
                                                           font=("Segoe UI", 10),
                                                           bg="#f7fafc", fg=Palette.TEXT_PRIMARY,
                                                           insertbackground=Palette.TEXT_PRIMARY,
                                                           relief="flat", padx=8, pady=8)
        self.txt_estadisticas.pack(fill="both", expand=True)
        self.txt_estadisticas.insert(tk.END, "Ejecute una validación para generar estadísticas.\n")

    def _crear_tree_en_notebook(self, notebook, titulo):
        frm = tk.Frame(notebook, bg=Palette.BG_PRIMARY)
        notebook.add(frm, text=f"   {titulo}   ")
        tree = ttk.Treeview(frm, style="Custom.Treeview")
        vsb = ttk.Scrollbar(frm, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frm.grid_rowconfigure(0, weight=1)
        frm.grid_columnconfigure(0, weight=1)
        return tree

    # ──────────────────────────────────────────────────────────────────────────
    # LÓGICA DE CARGA DE ARCHIVOS
    # ──────────────────────────────────────────────────────────────────────────
    def _verificar_dependencias(self):
        faltan = []
        if not PANDAS_OK:
            faltan.append("pandas / openpyxl")
        if not PDF_OK:
            faltan.append("PyPDF2 / pdfplumber")
        if not FUZZY_OK:
            faltan.append("fuzzywuzzy / python-Levenshtein")
        if faltan:
            messagebox.showwarning("Dependencias faltantes",
                "Instale:\n  pip install " + " ".join(faltan))

    def _cargar_archivo(self, es_origen: bool):
        ruta = filedialog.askopenfilename(
            title="Seleccionar archivo maestro",
            filetypes=[("Soportados", "*.xlsx *.xls *.pdf"),
                       ("Excel", "*.xlsx *.xls"),
                       ("PDF Tabular", "*.pdf"),
                       ("Todos", "*.*")]
        )
        if not ruta:
            return
        try:
            ext = Path(ruta).suffix.lower()
            if ext in (".xlsx", ".xls"):
                df = self.motor_archivos.read_excel(ruta)
            elif ext == ".pdf":
                df = self.motor_archivos.read_pdf_tablas(ruta)
                if df is None:
                    messagebox.showwarning("Sin tablas",
                        "No se detectaron tablas en el PDF.\nUse la pestaña de Planos para OCR.")
                    return
            else:
                messagebox.showerror("Formato inválido", "Solo Excel o PDF.")
                return

            if es_origen:
                self.df_origen = df
                self.lbl_archivo_origen.config(text=Path(ruta).name, fg=Palette.SUCCESS)
                self._llenar_combo(self.cmb_origen, list(df.columns))
                self._llenar_tree(self.tree_vista_origen, df)
                self._estado(f"Archivo 1 cargado: {len(df)} filas × {len(df.columns)} columnas")
            else:
                self.df_destino = df
                self.lbl_archivo_destino.config(text=Path(ruta).name, fg=Palette.SUCCESS)
                self._llenar_combo(self.cmb_destino, list(df.columns))
                self._llenar_tree(self.tree_vista_destino, df)
                self._estado(f"Archivo 2 cargado: {len(df)} filas × {len(df.columns)} columnas")
            logger.info(f"Cargado: {ruta}")
        except Exception as e:
            logger.error(traceback.format_exc())
            messagebox.showerror("Error de carga", str(e))

    def _cargar_origen(self):
        self._cargar_archivo(True)

    def _cargar_destino(self):
        self._cargar_archivo(False)

    def _llenar_combo(self, combo, columnas):
        combo["values"] = columnas
        if columnas:
            combo.current(0)

    def _llenar_tree(self, tree, df, max_filas=150):
        tree.delete(*tree.get_children())
        tree["columns"] = list(df.columns)
        tree["show"] = "headings"
        for col in df.columns:
            tree.heading(col, text=col)
            tree.column(col, width=110, anchor="w")
        for _, fila in df.head(max_filas).iterrows():
            vals = [str(v) if pd.notna(v) else "" for v in fila.values]
            tree.insert("", "end", values=vals)

    # ──────────────────────────────────────────────────────────────────────────
    # VALIDACIÓN CRUZADA
    # ──────────────────────────────────────────────────────────────────────────
    def _ejecutar_validacion(self):
        if self.df_origen is None or self.df_destino is None:
            messagebox.showwarning("Faltan archivos", "Cargue ambos archivos antes de continuar.")
            return
        col_o = self.var_col_origen.get()
        col_d = self.var_col_destino.get()
        if not col_o or not col_d:
            messagebox.showwarning("Faltan columnas", "Seleccione las columnas a comparar.")
            return
        try:
            self._estado("Ejecutando validación…")
            self.root.update()
            self.matcher = SparePartMatcher(
                umbral=self.var_umbral.get(),
                max_palabras=self.var_max_pal.get(),
                estrategia=self.var_estrategia.get()
            )
            def prog(act, tot):
                self.progreso["value"] = (act / tot) * 100
                self._estado(f"Progreso: {act}/{tot}")
                self.root.update_idletasks()
            self.df_resultado = self.matcher.comparar_dataframes(
                self.df_origen, col_o, self.df_destino, col_d, prog)
            self.progreso["value"] = 100
            self._llenar_tree(self.tree_resultados, self.df_resultado)
            if "COINCIDE" in self.df_resultado.columns:
                self._llenar_tree(self.tree_no_match,
                                  self.df_resultado[self.df_resultado["COINCIDE"] == "NO"])
            self.df_unificado = self.consolidador.consolidar(
                self.df_origen, self.df_destino, self.df_resultado, col_o, col_d)
            metricas = self.stats.analizar(self.df_resultado, self.df_origen, self.df_destino)
            self._mostrar_estadisticas(metricas)
            messagebox.showinfo("Validación completada",
                f"Coincidencias: {metricas['coincidencias']}\n"
                f"No coincidencias: {metricas['no_coincidencias']}\n"
                f"Tasa éxito: {metricas['tasa_exito']}%")
            logger.info("Validación finalizada exitosamente.")
        except Exception as e:
            logger.error(traceback.format_exc())
            messagebox.showerror("Error", str(e))
        finally:
            self._estado("Listo")

    def _mostrar_estadisticas(self, m):
        self.txt_estadisticas.delete("1.0", tk.END)
        texto = f"""╔══════════════════════════════════════════════════════════════════╗
║          REPORTE ESTADÍSTICO  —  SIVREP v{CONFIG.VERSION:8s}          ║
║          Generado: {Utils.ahora():40s}   ║
╠══════════════════════════════════════════════════════════════════╣

  MÉTRICAS GENERALES
  ─────────────────────────────────────────────────────────────────
  Total Archivo 1 (Origen):        {m['total_origen']:>8,}
  Total Archivo 2 (Destino):       {m['total_destino']:>8,}
  Registros Comparados:            {m['comparados']:>8,}
  Coincidencias Encontradas:       {m['coincidencias']:>8,}
  No Coincidencias:                {m['no_coincidencias']:>8,}

  INDICADORES DE DESEMPEÑO
  ─────────────────────────────────────────────────────────────────
  Tasa de Éxito:                   {m['tasa_exito']:>8.2f} %
  Cobertura sobre Origen:          {m['cobertura']:>8.2f} %
  Score Promedio (coincidencias):  {m['score_promedio']:>8.2f}
  Score Mínimo:                    {m['score_min']:>8.2f}
  Score Máximo:                    {m['score_max']:>8.2f}

  DISTRIBUCIÓN DE MÉTODOS
  ─────────────────────────────────────────────────────────────────
"""
        if "metodos" in m:
            for metodo, cantidad in m["metodos"].items():
                texto += f"  {metodo:35s}: {cantidad:>6,}\n"
        texto += "\n╚══════════════════════════════════════════════════════════════════╝\n"
        self.txt_estadisticas.insert("1.0", texto)

    def _exportar_resultados(self):
        if self.df_resultado is None:
            messagebox.showwarning("Sin datos", "Ejecute una validación primero.")
            return
        try:
            ruta = self.stats.generar_excel(self.df_resultado, self.df_unificado)
            messagebox.showinfo("Exportación exitosa", f"Reporte guardado en:\n{ruta}")
            self._estado(f"Excel exportado: {Path(ruta).name}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _exportar_unificado(self):
        if self.df_unificado is None:
            messagebox.showwarning("Sin datos", "Ejecute una validación primero.")
            return
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta = Path(CONFIG.OUTPUT_DIR) / f"LISTADO_UNIFICADO_{ts}.xlsx"
            self.df_unificado.to_excel(ruta, index=False, sheet_name="UNIFICADO")
            self.stats._formatear_excel(str(ruta))
            messagebox.showinfo("Exportación exitosa", f"Listado unificado:\n{ruta}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ──────────────────────────────────────────────────────────────────────────
    # PLANOS ELÉCTRICOS — OCR
    # ──────────────────────────────────────────────────────────────────────────
    def _cargar_plano(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar Plano Eléctrico",
            filetypes=[("PDF", "*.pdf"), ("Imágenes", "*.png *.jpg *.jpeg *.tiff *.bmp"), ("Todos", "*.*")]
        )
        if ruta:
            self.ruta_plano = ruta
            self.lbl_plano.config(text=Path(ruta).name, fg=Palette.SUCCESS)
            self._log_ocr(f"[INFO] Plano cargado: {ruta}")
            logger.info(f"Plano cargado: {ruta}")

    def _ejecutar_ocr(self):
        if not hasattr(self, "ruta_plano"):
            messagebox.showwarning("Sin archivo", "Seleccione un plano eléctrico.")
            return
        if not TESSERACT_OK:
            messagebox.showerror("OCR no disponible",
                "Instale pytesseract y el motor Tesseract-OCR en su sistema.")
            return
        if not PYMUPDF_OK and Path(self.ruta_plano).suffix.lower() == ".pdf":
            messagebox.showerror("Dependencia faltante",
                "Para procesar PDFs instale PyMuPDF:\n  pip install PyMuPDF\n\n"
                "Alternativa: convierta el PDF a imágenes PNG e intente de nuevo.")
            return
        try:
            self._log_ocr("[INFO] Iniciando procesamiento OCR…")
            self.root.update()
            ext = Path(self.ruta_plano).suffix.lower()
            if ext == ".pdf":
                resultados = self.motor_ocr.procesar_pdf(self.ruta_plano)
            else:
                resultados = self.motor_ocr.procesar_imagen(self.ruta_plano)
            self.df_ocr = pd.DataFrame(resultados)
            self._llenar_tree(self.tree_ocr, self.df_ocr)
            self._log_ocr(f"[ÉXITO] {len(resultados)} componentes detectados.")
            for r in resultados[:30]:
                self._log_ocr(f"  → TAG: {r['TAG']:<22} | MODELO: {r['MODELO']}")
            messagebox.showinfo("OCR completado", f"Componentes detectados: {len(resultados)}")
        except Exception as e:
            logger.error(traceback.format_exc())
            self._log_ocr(f"[ERROR] {str(e)}")
            messagebox.showerror("Error OCR", str(e))

    def _exportar_ocr(self):
        if not hasattr(self, "df_ocr") or self.df_ocr is None:
            messagebox.showwarning("Sin datos", "Procese un plano primero.")
            return
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta = Path(CONFIG.OUTPUT_DIR) / f"OCR_PLANOS_{ts}.xlsx"
            self.df_ocr.to_excel(ruta, index=False, sheet_name="COMPONENTES")
            self.stats._formatear_excel(str(ruta))
            messagebox.showinfo("Exportación exitosa", f"Resultados OCR:\n{ruta}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _log_ocr(self, mensaje):
        self.txt_log_ocr.config(state="normal")
        self.txt_log_ocr.insert(tk.END, mensaje + "\n")
        self.txt_log_ocr.see(tk.END)
        self.txt_log_ocr.config(state="disabled")

    # ──────────────────────────────────────────────────────────────────────────
    # ESTADÍSTICAS
    # ──────────────────────────────────────────────────────────────────────────
    def _generar_estadisticas(self):
        if self.df_resultado is None:
            messagebox.showwarning("Sin datos", "Ejecute una validación primero.")
            return
        try:
            m = self.stats.analizar(self.df_resultado, self.df_origen, self.df_destino)
            self._mostrar_estadisticas(m)
            self.notebook.select(self.tab_est)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _ver_graficos(self):
        if not MATPLOTLIB_OK:
            messagebox.showerror("Matplotlib no disponible", "Instale matplotlib.")
            return
        ruta = self.stats.generar_graficos()
        if ruta and Path(ruta).exists():
            ventana = tk.Toplevel(self.root)
            ventana.title("Gráficos Estadísticos")
            ventana.geometry("950x750")
            ventana.configure(bg=Palette.BG_PRIMARY)
            img = Image.open(ruta)
            img.thumbnail((900, 680))
            photo = ImageTk.PhotoImage(img)
            lbl = tk.Label(ventana, image=photo, bg=Palette.BG_PRIMARY)
            lbl.image = photo
            lbl.pack(padx=12, pady=12)
            tk.Button(ventana, text="Guardar imagen…", command=lambda: self._guardar_imagen(ruta),
                      bg=Palette.PRIMARY, fg=Palette.TEXT_ON_DARK, font=("Segoe UI", 10, "bold"),
                      relief="flat", cursor="hand2", padx=12, pady=5).pack(pady=8)

    def _ver_json(self):
        if not self.stats.reporte:
            messagebox.showwarning("Sin datos", "Genere estadísticas primero.")
            return
        ventana = tk.Toplevel(self.root)
        ventana.title("Reporte JSON")
        ventana.geometry("750x550")
        ventana.configure(bg=Palette.BG_PRIMARY)
        txt = scrolledtext.ScrolledText(ventana, wrap=tk.WORD, font=("Consolas", 10),
                                        bg="#f7fafc", fg=Palette.TEXT_PRIMARY, padx=8, pady=8)
        txt.pack(fill="both", expand=True, padx=10, pady=10)
        txt.insert("1.0", json.dumps(self.stats.reporte, indent=2, ensure_ascii=False, default=str))

    def _guardar_imagen(self, origen):
        destino = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG", "*.png")])
        if destino:
            shutil.copy(origen, destino)
            messagebox.showinfo("Guardado", f"Imagen guardada en:\n{destino}")

    def _estado(self, texto):
        self.barra_estado.config(text=f"  {texto}  ")


# ═══════════════════════════════════════════════════════════════════════════════
# PUNTO DE ENTRADA
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    root = tk.Tk()
    try:
        if sys.platform.startswith("win"):
            root.iconbitmap("sivrep.ico")
    except Exception:
        pass
    app = SIVREPApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
